import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# JSON-Datei einlesen
with open('archive_header.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# Neue Excel-Arbeitsmappe erstellen
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Link Paare"

# Überschriften in die erste Zeile einfügen
ws['A1'] = "Original URL"
ws['B1'] = "Archivierte URL"

# Zellen formatieren (blau und unterstrichen)
font_style = Font(color="0000FF", underline="single")

# Link-Daten einfügen
row = 2
for key, value in data.items():
    original_url = value['original_url']
    archivierte_url = value['archivierte_url']

    # Füge die URLs als Hyperlinks ein
    ws[f'A{row}'] = original_url
    ws[f'A{row}'].font = font_style
    ws[f'A{row}'].hyperlink = original_url

    ws[f'B{row}'] = archivierte_url
    ws[f'B{row}'].font = font_style
    ws[f'B{row}'].hyperlink = archivierte_url

    row += 1

# Spaltenbreite automatisch anpassen
for col in range(1, 3):
    col_letter = get_column_letter(col)
    ws.column_dimensions[col_letter].auto_size = True

# Excel-Datei speichern
wb.save("URLs.xlsx")

print("Excel-Datei wurde erfolgreich gespeichert!")
