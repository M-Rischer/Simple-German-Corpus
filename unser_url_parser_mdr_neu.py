import os
import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import defaultvalues as dfv

json_path = dfv.mdr_location + "parsed_header2025.json"

def parse_json_files(output_excel):
    data_pairs = []  # Use a list to maintain the order of pairs
        
    # Open and parse the JSON file
    with open(json_path, 'r') as f:
        try:
            json_data = json.load(f)
            
            for entry_key, entry_value in json_data.items():
                url = entry_value.get("url")
                easy = entry_value.get("easy", False)  # Default to False if not present
                matching_files = entry_value.get("matching_files", [])

                # Only proceed if 'easy' is True
                if not easy:
                    continue

                url_parts = url.split("/", 8)
                base_url = "/".join(url_parts[:7]) 

                for match in matching_files:
                    # Replace "__" with "/" in matching_files
                    formatted_match = match.replace("__", "/")
                    matching_url = f"https://www.{formatted_match}"
                    if matching_url.endswith("_normal.html"):
                        matching_url = matching_url[:-12]  # Remove "_normal.html"
                    elif url.endswith(".html"):
                        pass
                    else:
                        matching_url = matching_url[:-5]
                    
                    # Add the pair to the list in the order they are found
                    data_pairs.append((url, matching_url))  

        except json.JSONDecodeError:
            print(f"Error decoding JSON in file: {json_path}")

    # Create Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Parsed Data"

    # Write header row with bold font
    headers = [
        "Leichte Sprache", 
        "Standardsprache"
    ]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Write data rows in the order they were added
    for row_num, (source_url, matching_url) in enumerate(data_pairs, start=2):
        sheet.cell(row=row_num, column=1, value=source_url)
        sheet.cell(row=row_num, column=2, value=matching_url)

        # Make URLs clickable
        sheet.cell(row=row_num, column=1).hyperlink = source_url
        sheet.cell(row=row_num, column=2).hyperlink = matching_url

    # Adjust column widths
    for col_num, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = 50

    # Save the Excel file
    workbook.save(output_excel)

if __name__ == "__main__":
    output_excel = "urls_2025_LS_StS.xlsx"
    parse_json_files(output_excel)
    print(f"Excel file '{output_excel}' created successfully.")
