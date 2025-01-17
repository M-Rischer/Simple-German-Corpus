import os
import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def parse_json_files(root_folder, output_excel):
    data_pairs = []  # Use a list to maintain the order of pairs

    # Walk through directories
    for subdir, _, files in os.walk(root_folder):
        if "parsed_header.json" in files:
            json_path = os.path.join(subdir, "parsed_header.json")
            
            # Open and parse the JSON file
            with open(json_path, 'r') as f:
                try:
                    json_data = json.load(f)
                    
                    for entry_key, entry_value in json_data.items():
                        url = entry_value.get("url")
                        matching_files = entry_value.get("matching_files", [])
                        entry_type = entry_value.get("type")

                        # Only process entries with type "LS"
                        if entry_type == "LS" and url:
                            url_parts = url.split("/", 8)
                            base_url = "/".join(url_parts[:7]) 

                            for match in matching_files:
                                # Replace "__" with "/" in matching_files
                                formatted_match = match.replace("__", "/")
                                matching_url = f"{base_url}/{formatted_match}"
                                if matching_url.endswith("_normal.html"):
                                    matching_url = matching_url[:-12]  # Remove the last 12 characters "_normal.html"
                                elif url.endswith(".html"):
                                    pass
                                else:
                                    matching_url = matching_url[:-5]
                                
                                # Add the pair to the list in the order they are found
                                data_pairs.append((url, matching_url))  

                except json.JSONDecodeError:
                    print(f"Error decoding JSON in file: {json_path}")

    # Create Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Parsed Data"

    # Write header row with bold font
    headers = ["Texte in Leichter Sprache aus dem Simple German Corpus", "Texte in Alltagssprache aus dem Simple German Corpus"]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Write data rows in the order they were added
    for row_num, (source_url, matching_url) in enumerate(data_pairs, start=2):
        sheet.cell(row=row_num, column=1, value=source_url)
        sheet.cell(row=row_num, column=2, value=matching_url)

        # Make URLs clickable
        sheet.cell(row=row_num, column=1).hyperlink = source_url
        sheet.cell(row=row_num, column=2).hyperlink = matching_url

    # Adjust column widths
    for col_num, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = 50

    # Save the Excel file
    workbook.save(output_excel)

if __name__ == "__main__":
    root_folder = "C:/Users/Rischer/OneDrive - Helmut-Schmidt-Universität/Dokumente/GitLab/Simple-German-Corpus"
    output_excel = "urls.xlsx"
    parse_json_files(root_folder, output_excel)
    print(f"Excel file '{output_excel}' created successfully.")
