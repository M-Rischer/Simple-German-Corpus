import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import defaultvalues as dfv

what = ["www.lebenshilfe-main-taunus.de", "www.mdr.de"]
what_abv = ["lmt", "mdr"]

for i in range(len(what)):

    json_output = {}  # Dictionary für die JSON-Ausgabe 

    # JSON-Datei einlesen
    with open(f'{dfv.dataset_location}/{what[i]}/parsed_header.json', 'r', encoding='utf-8') as f:
        data = json.load(f)

    # Neue Excel-Arbeitsmappe erstellen
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Link Paare"

    # Überschriften in die erste Zeile einfügen
    ws['A1'] = "Original URL"
    ws['B1'] = "Archivierte URL"

    # Zellen formatieren (blau und unterstrichen)
    font_style = Font(color="0000FF", underline="single")

    # Link-Daten einfügen
    row = 2
    for key, value in data.items():
        archivierte_url = value['url']
        original_url = archivierte_url[42:]

        # Füge die URLs als Hyperlinks ein
        ws[f'A{row}'] = original_url
        ws[f'A{row}'].font = font_style
        ws[f'A{row}'].hyperlink = original_url

        ws[f'B{row}'] = archivierte_url
        ws[f'B{row}'].font = font_style
        ws[f'B{row}'].hyperlink = archivierte_url

        # Daten zur JSON-Struktur hinzufügen
        json_output[original_url] = {
            "archivierte_url": archivierte_url,
            "original_url": original_url
        }

        row += 1

    # Spaltenbreite automatisch anpassen
    for col in range(1, 3):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].auto_size = True

    # Excel-Datei speichern
    wb.save(f"Original_Archivierte_URL-Paare_{what_abv[i]}.xlsx")

    print("Excel-Datei wurde erfolgreich gespeichert!")

    # JSON-Datei speichern
    json_filename = f"Original_Archivierte_URL-Paare_{what_abv[i]}.json"
    with open(json_filename, 'w', encoding='utf-8') as json_file:
        json.dump(json_output, json_file, indent=4, ensure_ascii=False)
    print(f"JSON-Datei '{json_filename}' wurde erfolgreich gespeichert!")