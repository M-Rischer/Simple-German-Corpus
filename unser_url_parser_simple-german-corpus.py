import os
import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import defaultvalues as dfv

def parse_json_files(root_folder, output_excel):
    data_pairs = []  # Use a list to maintain the order of pairs
    all_entries = {}

    # Walk through directories and load all JSON entries
    for subdir, _, files in os.walk(root_folder):
        if "parsed_header.json" in files:
            json_path = os.path.join(subdir, "parsed_header.json")
            with open(json_path, 'r') as f:
                try:
                    json_data = json.load(f)
                    all_entries.update(json_data)  # Store all entries for lookup
                except json.JSONDecodeError:
                    print(f"Error decoding JSON in file: {json_path}")

    # Process the entries to match URLs
    for entry_key, entry_value in all_entries.items():
        url = entry_value.get("url")
        matching_files = entry_value.get("matching_files", [])
        entry_type = entry_value.get("type")

        # Only process entries with type "LS"
        if entry_type == "LS" and url:
            for match in matching_files:
                if match in all_entries:
                    matching_url = all_entries[match].get("url", "")
                    if matching_url:
                        data_pairs.append((url, matching_url))

    # Create Excel workbook and sheet
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Parsed Data"

    # Write header row with bold font
    headers = ["Texte in Leichter Sprache aus dem Simple German Corpus", "Texte in Alltagssprache aus dem Simple German Corpus"]
    for col_num, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)

    # Write data rows in the order they were added
    for row_num, (source_url, matching_url) in enumerate(data_pairs, start=2):
        sheet.cell(row=row_num, column=1, value=source_url)
        sheet.cell(row=row_num, column=2, value=matching_url)

        # Make URLs clickable
        sheet.cell(row=row_num, column=1).hyperlink = source_url
        sheet.cell(row=row_num, column=2).hyperlink = matching_url

    # Adjust column widths
    for col_num, _ in enumerate(headers, start=1):
        col_letter = get_column_letter(col_num)
        sheet.column_dimensions[col_letter].width = 50

    # Save the Excel file
    workbook.save(output_excel)

if __name__ == "__main__":
    root_folder = dfv.repository_location
    output_excel = "urls_sgc_ls_sts.xlsx"
    parse_json_files(root_folder, output_excel)
    print(f"Excel file '{output_excel}' created successfully.")
