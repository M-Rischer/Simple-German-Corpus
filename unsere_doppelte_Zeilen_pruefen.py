import openpyxl
from urllib.parse import urlparse
import re

# Funktion, um zu überprüfen, ob ein Wert ein Link ist
def is_link(value):
    if isinstance(value, str):
        return bool(re.match(r'http[s]?://', value))
    return False

# Funktion, um die vollständige URL (inkl. Endung) zu extrahieren
def get_full_url(value):
    parsed_url = urlparse(value.strip())
    return parsed_url.geturl()

# Entferne alle Hyperlinks aus einem Arbeitsblatt
def remove_all_hyperlinks(ws, url_columns):
    for col_idx in url_columns:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if cell.hyperlink:
                cell.hyperlink = None  # Entfernt die Verlinkung
                cell.font = openpyxl.styles.Font(color="000000", underline=None)  # Entfernt das Blau und die Unterstreichung

# Füge Hyperlinks wieder hinzu
def add_hyperlinks(ws, url_columns):
    for col_idx in url_columns:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            if is_link(cell.value):
                cell.hyperlink = cell.value  # Setzt die URL als Hyperlink
                cell.font = openpyxl.styles.Font(color="0000FF", underline="single")  # Blau und unterstrichen

# Lade die Excel-Datei und prüfe Duplikate
def remove_duplicates(file_path, url_columns):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Schritt 1: Entferne alle Hyperlinks
    print("Entferne alle Hyperlinks...")
    remove_all_hyperlinks(ws, url_columns)

    # Schritt 2: Prüfe Duplikate und lösche Zeilen
    print("Prüfe auf Duplikate...")
    duplicate_count = 0
    duplicates = {}
    rows_to_delete = set()
    original_rows = {}

    for col_idx in url_columns:
        print(f"Prüfe Spalte {col_idx}...")
        values_seen = set()

        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col_idx).value
            if is_link(cell_value):
                full_url = get_full_url(cell_value)

                if full_url in values_seen:
                    rows_to_delete.add(row)
                    duplicate_count += 1
                    if full_url not in duplicates:
                        duplicates[full_url] = []
                    duplicates[full_url].append(row)
                else:
                    values_seen.add(full_url)
                    original_rows[full_url] = row

    # Debugging: Zeige Duplikate an
    if duplicate_count > 0:
        print(f"Anzahl der Duplikate: {duplicate_count}")
        for value, rows in duplicates.items():
            original_row = original_rows.get(value)
            print(f"Duplikat-URL '{value}' gefunden in den Zeilen: {rows}")
            print(f"Original befindet sich in Zeile {original_row}")
    else:
        print("Keine doppelten Einträge gefunden.")

    # Lösche die Duplikate von unten nach oben
    print("Lösche Duplikate...")
    for row in sorted(rows_to_delete, reverse=True):
        print(f"Lösche Zeile {row}...")
        ws.delete_rows(row)

    # Schritt 3: Füge die Hyperlinks wieder hinzu
    print("Erstelle Hyperlinks neu...")
    add_hyperlinks(ws, url_columns)

    # Speichere die aktualisierte Datei
    updated_file_path = 'updated_' + file_path
    wb.save(updated_file_path)
    print(f"Duplikate entfernt. Aktualisierte Datei gespeichert unter: {updated_file_path}")

# Beispielaufruf der Funktion
remove_duplicates("urls_2025_LS_StS.xlsx", url_columns=[1, 2])
