import openpyxl
import re
from urllib.parse import urlparse

# Funktion, um zu überprüfen, ob ein Wert ein Link ist
def is_link(value):
    if isinstance(value, str):
        # Einfacher regulärer Ausdruck zur Erkennung von URLs (beginnend mit http:// oder https://)
        return bool(re.match(r'http[s]?://', value))
    return False

# Funktion, um die vollständige URL (inkl. Endung) zu extrahieren
def get_full_url(value):
    parsed_url = urlparse(value.strip())  # Entferne führende und nachfolgende Leerzeichen
    # Gib die vollständige URL ohne Modifikationen zurück
    return parsed_url.geturl()

# Lade die Excel-Datei und prüfe Duplikate
def remove_duplicates(file_path, url_columns):
    # Öffne die Excel-Datei
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active  # Wähle das aktive Arbeitsblatt aus
    
    # Zählvariablen für die Duplikate
    duplicate_count = 0
    duplicates = {}
    rows_to_delete = set()  # Set für Zeilen, die gelöscht werden sollen
    original_rows = {}  # Dictionary, um die erste Vorkommenszeile für jede URL zu speichern
    
    # Prüfe jede angegebene Spalte separat
    for col_idx in url_columns:
        values_seen = set()  # Set für bereits gesehene vollständige URLs in der aktuellen Spalte
        
        for row in range(2, ws.max_row + 1):  # Starte bei der zweiten Zeile (überspringe Kopfzeile)
            cell_value = ws.cell(row=row, column=col_idx).value
            
            if is_link(cell_value):  # Prüfe, ob der Zellinhalt ein Link ist
                full_url = get_full_url(cell_value)  # Verwende die vollständige URL inkl. Endung
                if full_url in values_seen:
                    rows_to_delete.add(row)  # Markiere die Zeile mit Duplikat zum Löschen
                    duplicate_count += 1
                    if full_url not in duplicates:
                        duplicates[full_url] = []
                    duplicates[full_url].append(row)
                else:
                    values_seen.add(full_url)
                    original_rows[full_url] = row  # Speichere die Originalzeile (erste Vorkommen)
    
    # Gib die Duplikate und deren Zeilen klar aus
    if duplicate_count > 0:
        print(f"Anzahl der Duplikate (vollständige URLs): {duplicate_count}")
        for value, rows in duplicates.items():
            original_row = original_rows.get(value)  # Hole die Originalzeile der URL
            print(f"Duplikat-URL '{value}' gefunden in den folgenden Zeilen: {rows}")
            print(f"Original-URL '{value}' befindet sich in Zeile {original_row}")
    else:
        print("Keine doppelten Einträge gefunden.")
    
    # Lösche die Duplikate, nachdem wir alle Spalten überprüft haben
    for row in sorted(rows_to_delete, reverse=True):  # Lösche Zeilen von unten nach oben
        ws.delete_rows(row)
    
    # Speichere die Datei mit den Änderungen
    updated_file_path = 'updated_' + file_path
    wb.save(updated_file_path)
    
    print(f"Duplikate entfernt. Aktualisierte Datei gespeichert unter: {updated_file_path}")

# Beispielaufruf der Funktion
remove_duplicates("URLs.xlsx", url_columns=[1, 2])
